"""Export an audit table of all news articles considered for a given newsletter window.

Reads the SQLite audit store and produces both a CSV and an XLSX with one row per
article. Each row carries:
  - Source (BSE / NewsAPI / RSS / RBI / NHB / etc.)
  - Company (the competitor it was tagged to, blank for policy feeds)
  - Title and Published date
  - URL (the canonical article / filing link)
  - LLM (provider + model that processed it at the relevant stage)
  - Newsletter section (best-effort mapping to the rendered index section)
  - Status (Kept / Dropped — with reason if dropped)

Run:
    python -m services.export_audit_table --week-start 2026-05-16 --week-end 2026-05-31
The two output files are written to:
    audit/audit_table_<start>_to_<end>.csv
    audit/audit_table_<start>_to_<end>.xlsx
"""
from __future__ import annotations

import argparse
import csv
import json
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import get_settings


# ----- helpers --------------------------------------------------------------------


def _bucket_from_bse_kept(payload: Dict[str, Any]) -> str:
    """Map a bse_kept payload to a fully-qualified newsletter index label."""
    bucket = (payload.get("bucket") or "").strip()
    return f"Competitor Intelligence — {bucket}" if bucket else "Competitor Intelligence"


def _latest_pipeline_for_article(
    conn: sqlite3.Connection, article_id: str, stage: str
) -> Optional[Dict[str, Any]]:
    row = conn.execute(
        """
        SELECT payload_json, created_at
          FROM article_pipeline
         WHERE article_id = ? AND stage = ?
      ORDER BY created_at DESC
         LIMIT 1
        """,
        (article_id, stage),
    ).fetchone()
    if not row:
        return None
    try:
        return json.loads(row[0])
    except (ValueError, TypeError):
        return None


def _llm_provider_model_for_stage(
    conn: sqlite3.Connection, stage: str
) -> Tuple[str, str]:
    """Pick the latest provider/model recorded for a stage. Empty if not audited."""
    row = conn.execute(
        """
        SELECT provider, model FROM llm_system_instruction_audit
         WHERE stage = ?
      ORDER BY created_at DESC LIMIT 1
        """,
        (stage,),
    ).fetchone()
    if not row:
        return "", ""
    return (row[0] or "", row[1] or "")


def _classify_source(source: str, url: str) -> str:
    """Group long-tail news source strings into the four ingestion lanes."""
    s = (source or "").strip()
    if not s:
        # Aggregator-routed articles sometimes have an empty source string.
        return "News (uncategorised)"
    if s == "BSE":
        return "BSE"
    if s in {"RBI", "NHB"}:
        return s
    if "google.com" in url.lower() or "news.google.com" in url.lower():
        return "RSS / Google News"
    # Everything else is a NewsAPI / RSS publisher.
    return f"News ({s})"


# ----- main builder ---------------------------------------------------------------


HEADERS = [
    "#",
    "Source",
    "Company",
    "Title",
    "Published",
    "URL",
    "LLM Extract",
    "Newsletter Section (index)",
    "Status",
    "Notes",
]


def build_rows(db_path: str, start: str, end: str) -> List[List[Any]]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    # Articles whose published_at falls within the window. Empty published_at strings
    # are excluded to avoid noisy aggregator entries with no usable date.
    articles = conn.execute(
        """
        SELECT article_id, url, company, title, source, published_at
          FROM articles
         WHERE published_at >= ? AND published_at <= ?
      ORDER BY source, company, published_at
        """,
        (start, end),
    ).fetchall()

    # Pre-fetch LLM provider/model per stage we care about.
    llm_by_stage = {
        stage: _llm_provider_model_for_stage(conn, stage)
        for stage in (
            "bse_classification",
            "bse_pdf_enrichment",
            "bse_rating_enrichment",
            "curated_evidence_selection",
            "agentic_industry_output",
            "agentic_regulatory_output",
            "agentic_competitor_output",
            "newsletter_composer_output",
        )
    }

    out: List[List[Any]] = []
    for i, a in enumerate(articles, start=1):
        article_id = a["article_id"]
        source_label = _classify_source(a["source"], a["url"])

        section = ""
        status = ""
        notes = ""
        llm_stage = ""

        if a["source"] == "BSE":
            llm_stage = "bse_classification"
            kept = _latest_pipeline_for_article(conn, article_id, "bse_kept")
            if kept:
                section = _bucket_from_bse_kept(kept)
                status = "Kept"
                action = (kept.get("action") or "").strip()
                person = (kept.get("person_name") or "").strip()
                if person or action:
                    notes = " · ".join([p for p in (person, action, kept.get("event_date", "")) if p])
                rating_enrich = _latest_pipeline_for_article(conn, article_id, "bse_rating_enrichment")
                if rating_enrich and rating_enrich.get("success"):
                    # Surface that PDF rating enrichment ran on this filing.
                    notes = (notes + " · " if notes else "") + "rating-enriched"
                    llm_stage = "bse_rating_enrichment"
            else:
                dropped = _latest_pipeline_for_article(conn, article_id, "bse_gate_dropped")
                classify = _latest_pipeline_for_article(conn, article_id, "bse_classification")
                if dropped:
                    section = "(dropped)"
                    status = "Dropped — Operational gate"
                    notes = dropped.get("reason", "missing_required_field")
                elif classify:
                    cat = (classify.get("category") or "").strip()
                    if cat in ("", "Other"):
                        section = "(dropped)"
                        status = "Dropped — Classified as Other"
                    else:
                        # Classified into a relevant category but no bse_kept event written
                        # — either materiality cap excluded it, or run was interrupted before
                        # the kept-write. Show the intended bucket so the auditor can see it.
                        section = f"(intended) {cat}"
                        status = "Dropped — Materiality cap"
                else:
                    # BSE classifier only audits relevant=true items. No event = the
                    # batch LLM classified this filing as "Other" (routine), so it was
                    # silently skipped without an individual audit row.
                    section = "(dropped)"
                    status = "Dropped — Routine / Other"
        else:
            # News / RSS / policy lane. The curated_evidence_selection event indicates
            # an article was selected for per-competitor LLM synthesis.
            curated = _latest_pipeline_for_article(conn, article_id, "curated_evidence_selection")
            scoring = _latest_pipeline_for_article(conn, article_id, "scoring")
            llm_signal = _latest_pipeline_for_article(conn, article_id, "llm_signal")

            if a["source"] in ("RBI", "NHB"):
                section = "Industry Pulse / Regulatory Watch (policy feed)"
                status = "Kept" if curated else "Considered"
                llm_stage = "agentic_regulatory_output"
            elif curated:
                section = "Competitor Intelligence (via news)"
                status = "Selected for synthesis"
                llm_stage = "curated_evidence_selection"
            elif llm_signal:
                section = "Competitor Intelligence (via news)"
                status = "Signal-tagged but not selected"
                llm_stage = "llm_signal"
            elif scoring:
                section = "(dropped)"
                status = "Scored but not selected"
            else:
                section = "(dropped)"
                status = "Fetched, not processed"

        provider, model = llm_by_stage.get(llm_stage, ("", ""))
        llm_label = f"{provider} / {model}" if (provider or model) else ""

        out.append([
            i,
            source_label,
            a["company"] or "",
            a["title"] or "",
            a["published_at"] or "",
            a["url"] or "",
            llm_label,
            section,
            status,
            notes,
        ])

    conn.close()
    return out


# ----- writers --------------------------------------------------------------------


def write_csv(rows: List[List[Any]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(HEADERS)
        w.writerows(rows)


def write_xlsx(rows: List[List[Any]], out_path: Path, start: str, end: str) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Article Audit"

    title_font = Font(bold=True, size=13)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F2D63")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.cell(row=1, column=1, value=f"Article audit — {start} to {end}").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")

    for r_idx, row in enumerate(rows, start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = wrap
            if HEADERS[c_idx - 1] == "URL" and value:
                cell.hyperlink = str(value)
                cell.font = Font(color="0F4DBF", underline="single")

    widths = {
        "#": 6,
        "Source": 18,
        "Company": 22,
        "Title": 60,
        "Published": 12,
        "URL": 50,
        "LLM Extract": 22,
        "Newsletter Section (index)": 32,
        "Status": 26,
        "Notes": 30,
    }
    for col_idx, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 18)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}{3 + len(rows)}"
    wb.save(out_path)


# ----- CLI ------------------------------------------------------------------------


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--week-start", required=True)
    ap.add_argument("--week-end", required=True)
    ap.add_argument("--out-dir", default="audit")
    args = ap.parse_args()

    db = get_settings().sqlite_db_path
    rows = build_rows(db, args.week_start, args.week_end)

    out_dir = Path(args.out_dir)
    stem = f"audit_table_{args.week_start}_to_{args.week_end}"
    csv_path = out_dir / f"{stem}.csv"
    xlsx_path = out_dir / f"{stem}.xlsx"

    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path, args.week_start, args.week_end)

    # Summary line for quick triage.
    by_status: Dict[str, int] = {}
    for r in rows:
        by_status[r[8]] = by_status.get(r[8], 0) + 1
    print(f"Articles in [{args.week_start} … {args.week_end}]: {len(rows)}")
    for status, n in sorted(by_status.items(), key=lambda kv: -kv[1]):
        print(f"  {status or '(blank)':40s} {n}")
    print(f"\nWrote {csv_path}")
    print(f"Wrote {xlsx_path}")


if __name__ == "__main__":
    main()
